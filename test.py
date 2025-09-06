import pyvisa
import time
from openpyxl import load_workbook
#wb = load_workbook('data.xlsx')
#sheet=wb.actice
ip_address="169.254.189.200"
power =-10
def connect_znb8(ip_address,power):
    """
    Kết nối tới máy ZNB8 qua LAN

    Parameters:
    ip_address (str): Địa chỉ IP của máy ZNB8

    Returns:
    rm: Resource Manager
    inst: Thiết bị ZNB8 đã kết nối
    """
    global inst
    try:

        # Khởi tạo Resource Manager
        rm = pyvisa.ResourceManager()
        # Tạo chuỗi kết nối VISA
        visa_address = f"TCPIP0::{ip_address}::inst0::INSTR"
        # Mở kết nối
        inst = rm.open_resource(visa_address)
        #inst.write("CALC1:PAR:SDEF 'Trc2', 'S11'")
        inst.write('SENS:SWE:TYPE POIN')
        inst.write('SENS:SWE:POIN 201')
        #inst.write('FREQ:CW {0}'.format(frequency))
        inst.write('SOUR:POW3 {0}'.format(power))
        inst.write('OUTP ON')
        inst.timeout = 10000  # Thiết lập timeout 10 giây


        return rm, inst

    except Exception as e:
        print(f"Lỗi khi kết nối: {str(e)}")
        return None, None
def connect_powermeter():
    rm = pyvisa.ResourceManager()
    address = "GPIB2::13::INSTR"
    instru = rm.open_resource(address)
    instru.timeout = 10000
    instru.write("FREQ {0}".format(frequency))
    instru.write("INIT:IMM")
    time.sleep(1.5)
    a = round(float(instru.query("FETCH:POW?")), 2)
    return a
from openpyxl import load_workbook
file_path = r"C:\Users\Duyhang\OneDrive - Hanoi University of Science and Technology\Desktop\data.xlsx"
wb = load_workbook(file_path)
sheet = wb.active
x=1
connect_znb8(ip_address,power)
while x<677:
    frequency = sheet.cell(row=x, column=3).value
    print(frequency)
    inst.write('FREQ:CW {0}'.format(frequency))
    sheet.cell(row =x,column =4).value=connect_powermeter()
    cell_value1 = sheet.cell(row=x, column=4).value
    print(cell_value1)
    wb.save(r"C:\Users\Duyhang\OneDrive - Hanoi University of Science and Technology\Desktop\data.xlsx")
    x=x+1
inst.write('OUTP OFF')

